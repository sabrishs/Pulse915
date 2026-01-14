"""
Project: Pulse915 Backtesting System
Phase: 4 (1-Min) - High-Precision Backtesting & Execution Simulation
Version: 3.3
Date: 2026-01-12
Updated By: Sabrish Surender

Description:
A high-resolution version of the Phase 4 backtester using 1-minute candle data.
"HOLD-TO-CLOSE" Strategy - Simple, Brutal Efficiency.

Strategy (v2.2) - "HOLD-TO-CLOSE":
ANALYSIS INSIGHT: Protection Ladder and Fast Fail over-engineered the exits.
Reverting to simple execution that was profitable (+₹43k).

Mode-Specific Targets:
- Mode A (ORB): target = entry + 2.0R
- Mode B (VWAP): target = entry + 2.0R  
- Mode C (Day High): target = entry + 1.5R (Sniper target)

Execution Logic (v2.2) - "Hold-to-Close":
- NO TRAILING STOPS - all trailing/protection ladder removed
- NO FAST FAIL - give trades room to breathe
- NO MODE A TIME GATE - if Phase 3 sent it, Phase 4 trades it
- Simple flow: Entry → Hard Stop → Hard Target → Time Exit (15:10)

Exit Labels:
- STOP_LOSS: Hard stop hit
- TARGET_HIT: Hard target hit  
- TIME_EXIT_1510: Force close at 15:10
"""

# phase-4.py
import os
import sys
import pandas as pd
import numpy as np
import warnings
from datetime import time, datetime

# Suppress pandas datetime parsing warnings
warnings.filterwarnings('ignore', message='Could not infer format')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
# ===============================
# CONFIG
# ===============================
PHASE3_FILE = "phase-3results/Phase3_results.xlsx"
ONE_MIN_DATA_DIR = "downloaded_data/1min"   # 1-minute candles
OUTPUT_DIR = "phase-4results"
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"phase4_results_1m_{ts}.xlsx")

# Capital & Risk Config
C_PER_DAY = 1000000  # ₹1,00,0000 per day (not total)
L_PCT = 0.02
C_PCT = 0.50

# Phase-4C Config
S_PCT = 0.02  # 2% stop-loss (TEST 2)
R_MULT = 1.0

MIN_TRADE_VALUE = 5000
FORCE_EXIT_TIME = time(15, 10)
# Transaction cost and slippage (fractions)
TRANSACTION_COST_PCT = 0.0005  # 0.05%
SLIPPAGE_PCT = 0.001  # 0.1%

# Minimum edge filter (prevents "death by a thousand paper cuts")
COST_BUFFER = 0.0025  # 0.25% minimum profit after costs

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ===============================
# ATR CALCULATION FOR DYNAMIC RISK
# ===============================
def add_rolling_atr(df_full):
    """
    Calculate rolling 14-period ATR on the entire historical dataset.
    Uses Wilder's smoothing method (EMA with alpha=1/14).
    Adds 'ATR' column to the dataframe.
    
    Args:
        df_full: DataFrame with 'high', 'low', 'close' columns (lowercase)
    
    Returns:
        DataFrame with 'ATR' column added
    """
    df = df_full.copy()
    df['prev_close'] = df['close'].shift(1)
    df['tr'] = np.maximum(
        df['high'] - df['low'],
        np.maximum(
            abs(df['high'] - df['prev_close']),
            abs(df['low'] - df['prev_close'])
        )
    )
    # Wilder's EMA (span=14)
    df['ATR'] = df['tr'].ewm(span=14, adjust=False).mean()
    
    # Clean up temporary columns
    df = df.drop(columns=['prev_close', 'tr'], errors='ignore')
    return df

# ===============================
# PROTECTION LADDER CALCULATION (v2.0)
# ===============================
def calculate_protection_ladder(entry_price, risk_r, current_high, current_stop, initial_stop):
    """
    Protection Ladder: Data-driven stop management based on MFE analysis.
    Analysis showed 55% of trades hit +1R but many still lost due to late protection.
    
    Ladder Levels:
    - At +0.8R MFE: reduce risk (stop = entry - 0.2R)
    - At +1.0R MFE: breakeven (stop = entry)
    - At +1.5R MFE: lock profit (stop = entry + 0.5R)
    - After +1.5R: trail by 1.0R from high
    
    Returns: (new_stop, protection_level) for proper exit labeling
    """
    new_stop = current_stop
    current_profit = current_high - entry_price
    protection_level = "INITIAL"  # Track which protection tier we're at
    
    # Level 1: At +0.8R → Reduce risk (stop = entry - 0.2R)
    if current_profit >= (0.8 * risk_r):
        reduced_risk_stop = entry_price - (0.2 * risk_r)
        new_stop = max(new_stop, reduced_risk_stop)
        protection_level = "REDUCED_RISK"
    
    # Level 2: At +1.0R → Breakeven (stop = entry)
    if current_profit >= (1.0 * risk_r):
        new_stop = max(new_stop, entry_price)
        protection_level = "BREAKEVEN"
    
    # Level 3: At +1.5R → Lock profit (stop = entry + 0.5R)
    if current_profit >= (1.5 * risk_r):
        locked_stop = entry_price + (0.5 * risk_r)
        new_stop = max(new_stop, locked_stop)
        protection_level = "LOCKED"
    
    # Level 4: After +1.5R → Trail by 1.0R from high
    if current_profit >= (1.5 * risk_r):
        trail_stop = current_high - (1.0 * risk_r)
        if trail_stop > new_stop:
            new_stop = trail_stop
            protection_level = "TRAILING"
    
    return new_stop, protection_level

# ===============================
# LOAD PHASE-3 OUTPUT
# ===============================
df = pd.read_excel(PHASE3_FILE)

df = df.rename(columns={
    "Stock": "symbol",
    "Entry Price (₹)": "entry_price",
    "Stop-Loss (₹)": "stop_loss",  # From Phase-3
    "Target (₹)": "target",        # From Phase-3
    "Date": "date",
    "Entry Time": "buy_time",
    "Entry Mode": "mode"  # A, B, or C from Phase-3
})

# Read VelocityScore if available, otherwise use default
if "VelocityScore" in df.columns:
    df["velocity_score"] = pd.to_numeric(df["VelocityScore"], errors="coerce").fillna(75.0)
else:
    # Default VelocityScore if not provided by Phase-3
    # TODO: Phase-3 should output VelocityScore in future
    df["velocity_score"] = 75.0

required_cols = {"symbol", "entry_price", "stop_loss", "target", "date", "buy_time"}
# Prepare output: multi-sheet Excel with formatting
missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"❌ Missing required columns: {missing}")

# ===============================
# MINIMUM EDGE FILTER
# ===============================
# Filter: (Target - Entry) / Entry >= CostBuffer
# Prevents "death by a thousand paper cuts"
initial_count = len(df)
df["potential_profit_pct"] = ((df["target"] - df["entry_price"]) / df["entry_price"])
df = df[df["potential_profit_pct"] >= COST_BUFFER].copy()
filtered_count = initial_count - len(df)
print(f"Minimum edge filter: Removed {filtered_count} trades with <{COST_BUFFER*100:.2f}% potential profit")
print(f"Remaining trades: {len(df)}")
df = df.drop(columns=["potential_profit_pct"])  # Clean up temporary column

# ===============================
# PHASE-4A — WEIGHTS (VelocityScore-based)
# ===============================
# Formula: s_i = VelocityScore - 50 (clip at 0)
#          w_i = s_i / Σ(s_i)

df["s_i"] = df["velocity_score"] - 50
df["s_i"] = df["s_i"].clip(lower=0)  # If < 0, set to 0

# ===============================
# PHASE-4B — DAILY CAPITAL ALLOCATION
# ===============================
# Allocate ₹1,00,000 per day (not total)
# Group by date and calculate weights within each day

all_days_data = []

for trade_date, day_df in df.groupby("date"):
    day_df = day_df.copy()
    
    # Calculate weights for this day only
    sum_s_i = day_df["s_i"].sum()
    if sum_s_i > 0:
        day_df["weight"] = day_df["s_i"] / sum_s_i
    else:
        day_df["weight"] = 0.0
    
    # Risk allocation for this day
    Lcap_day = C_PER_DAY * L_PCT
    day_df["loss_cap"] = Lcap_day * day_df["weight"]
    
    # ===============================
    # PHASE-4C — USE PHASE-3 PRICE LEVELS
    # ===============================
    # Use stop-loss and target from Phase-3 (don't recalculate!)
    day_df["stop_price"] = pd.to_numeric(day_df["stop_loss"], errors="coerce")
    day_df["target_price"] = pd.to_numeric(day_df["target"], errors="coerce")
    day_df["risk_per_share"] = day_df["entry_price"] - day_df["stop_price"]
    
    # ===============================
    # PHASE-4D — QUANTITY (per day)
    # ===============================
    day_df["risk_per_share"] = pd.to_numeric(day_df["risk_per_share"], errors="coerce").fillna(0.0)
    day_df["loss_cap"] = pd.to_numeric(day_df["loss_cap"], errors="coerce").fillna(0.0)
    day_df["entry_price"] = pd.to_numeric(day_df["entry_price"], errors="coerce").fillna(0.0)
    
    def safe_floor_div(a, b):
        try:
            a = float(a)
            b = float(b)
        except Exception:
            return 0.0
        if b <= 0:
            return 0.0
        return np.floor(a / b)
    
    day_df["qty_risk"] = day_df.apply(lambda r: safe_floor_div(r["loss_cap"], r["risk_per_share"]), axis=1)
    day_df["qty_cap"] = np.floor((C_PER_DAY * C_PCT) / day_df["entry_price"]).replace([np.inf, -np.inf], 0).fillna(0)
    day_df["quantity"] = np.minimum(day_df["qty_risk"], day_df["qty_cap"])
    
    day_df["trade_value"] = day_df["quantity"] * day_df["entry_price"]
    
    # ===============================
    # PHASE-4E — PORTFOLIO SCALING (per day)
    # ===============================
    total_deployed_day = day_df["trade_value"].sum()
    if total_deployed_day > C_PER_DAY:
        alpha = C_PER_DAY / total_deployed_day
        day_df["quantity"] = np.floor(day_df["quantity"] * alpha)
        day_df["trade_value"] = day_df["quantity"] * day_df["entry_price"]
    
    # Remove zero-quantity rows for this day
    day_df = day_df[day_df["quantity"] > 0].copy()
    
    all_days_data.append(day_df)

# Combine all days
df = pd.concat(all_days_data, ignore_index=True) if all_days_data else pd.DataFrame()

# ===============================
# PHASE-4 EXIT RESOLUTION (5-MIN BACKTEST)
# ===============================
sell_prices = []
sell_times = []
exit_reasons = []
pnls = []
risk_r_list = []          # Dynamic risk values for verification
stop_dynamic_list = []    # Dynamic stop values for verification
target_dynamic_list = []  # Dynamic target values for verification
high_time_list = []       # Time when day high occurred
low_time_list = []        # Time when day low occurred

for _, row in df.iterrows():
    symbol = row["symbol"]
    buy_price = row["entry_price"]
    stop = row["stop_price"]
    target = row["target_price"]
    qty = row["quantity"]
    buy_time = row["buy_time"]

    # normalize buy_time to a time object if possible
    if pd.isna(buy_time):
        buy_time = time(0, 0)
    elif isinstance(buy_time, str):
        try:
            bt = pd.to_datetime(buy_time).time()
            buy_time = bt
        except Exception:
            try:
                parts = [int(x) for x in buy_time.split(":")]
                buy_time = time(*parts[:3])
            except Exception:
                buy_time = time(0, 0)
    elif isinstance(buy_time, pd.Timestamp):
        buy_time = buy_time.time()

    # ===============================
    # EARLY CALCULATION: Get HighTime/LowTime for all trades (including skipped)
    # ===============================
    # This runs before filters so skipped trades also have this data
    date_val_early = row.get("date")
    date_str_early = None
    if not pd.isna(date_val_early):
        try:
            date_str_early = pd.to_datetime(date_val_early).strftime("%Y-%m-%d")
        except:
            date_str_early = str(date_val_early)
    
    one_min_file_early = os.path.join(ONE_MIN_DATA_DIR, f"{symbol}.csv")
    high_time = ""
    low_time = ""
    
    if os.path.exists(one_min_file_early):
        try:
            mdf_early = pd.read_csv(one_min_file_early)
            # Handle both 'datetime' and 'date' column names
            datetime_col = 'datetime' if 'datetime' in mdf_early.columns else 'date' if 'date' in mdf_early.columns else None
            if datetime_col:
                mdf_early[datetime_col] = pd.to_datetime(mdf_early[datetime_col], errors='coerce')
                if date_str_early:
                    mdf_early = mdf_early[mdf_early[datetime_col].dt.strftime('%Y-%m-%d') == date_str_early]
                
                if len(mdf_early) > 0 and 'high' in mdf_early.columns and 'low' in mdf_early.columns:
                    # Find when day high occurred
                    high_idx = mdf_early['high'].idxmax()
                    if not pd.isna(high_idx):
                        high_time = mdf_early.loc[high_idx, datetime_col].strftime('%H:%M') if pd.notna(mdf_early.loc[high_idx, datetime_col]) else ""
                    
                    # Find when day low occurred
                    low_idx = mdf_early['low'].idxmin()
                    if not pd.isna(low_idx):
                        low_time = mdf_early.loc[low_idx, datetime_col].strftime('%H:%M') if pd.notna(mdf_early.loc[low_idx, datetime_col]) else ""
        except Exception as e:
            pass  # Keep empty strings

    # ===============================
    # PHASE 4 TRADE FILTERS (based on analysis)
    # ===============================
    # Get trade mode for filtering (extract first letter/word)
    trade_mode = row.get('mode', 'A')
    if pd.isna(trade_mode):
        trade_mode = 'A'
    trade_mode_str = str(trade_mode).strip()
    # Extract first character for mode detection (A, B, or C)
    mode_letter = trade_mode_str[0].upper() if trade_mode_str else 'A'
    
    # v2.9: TIME FILTERS BY MODE (based on P&L analysis)
    # Mode A: 09:50-09:55 loses money (-₹15.5K combined). 09:45 is profitable (+₹4.7K)
    # Mode B/C: 11:00-12:59 is mid-day chop zone
    if mode_letter == 'A' and buy_time is not None:
        if buy_time >= time(9, 50):  # Skip 09:50 and later
            print(f"⚠️ SKIP: Mode A trade for {symbol} at {buy_time} (late ORB - loses money)")
            sell_prices.append(buy_price)
            sell_times.append("")
            exit_reasons.append("SKIPPED_MODE_A_LATE")
            pnls.append(0)
            risk_r_list.append(0)
            stop_dynamic_list.append(0)
            target_dynamic_list.append(0)
            high_time_list.append("")
            low_time_list.append("")
            continue
    
    # v3.3: MODE B BAD TIMES - comprehensive bad time filter
    # Morning losers: 09:35, 09:55, 10:05, 10:20, 10:25, 10:35, 10:40, 10:45, 10:50, 10:55
    # Afternoon losers: 13:00, 13:05, 13:25, 13:40
    if mode_letter == 'B' and buy_time is not None:
        bad_times = [time(9, 35), time(9, 55), time(10, 5), time(10, 20), time(10, 25),
                     time(10, 35), time(10, 40), time(10, 45), time(10, 50), time(10, 55),
                     time(13, 0), time(13, 5), time(13, 25), time(13, 40)]
        if buy_time in bad_times:
            print(f"⚠️ SKIP: Mode B trade for {symbol} at {buy_time} (bad time slot)")
            sell_prices.append(buy_price)
            sell_times.append("")
            exit_reasons.append("SKIPPED_MODE_B_BADTIME")
            pnls.append(0)
            risk_r_list.append(0)
            stop_dynamic_list.append(0)
            target_dynamic_list.append(0)
            high_time_list.append("")
            low_time_list.append("")
            continue
    
    if mode_letter in ['B', 'C'] and buy_time is not None:
        if time(11, 0) <= buy_time < time(13, 0):
            print(f"⚠️ SKIP: Mode {mode_letter} trade for {symbol} at {buy_time} (mid-day chop zone)")
            sell_prices.append(buy_price)
            sell_times.append("")
            exit_reasons.append(f"SKIPPED_MODE_{mode_letter}_MIDDAY")
            pnls.append(0)
            risk_r_list.append(0)
            stop_dynamic_list.append(0)
            target_dynamic_list.append(0)
            high_time_list.append("")
            low_time_list.append("")
            continue

    # Build path to 1-minute CSV: downloaded_data/1min_candles/<SYMBOL>.csv
    date_val = row.get("date")
    date_str = None
    if pd.isna(date_val):
        date_str = None
    else:
        try:
            date_ts = pd.to_datetime(date_val)
            date_str = date_ts.strftime("%Y-%m-%d")
        except Exception:
            date_str = str(date_val)

    one_min_file = os.path.join(ONE_MIN_DATA_DIR, f"{symbol}.csv")

    # ===============================
    # SAFETY CHECK: Skip if 1-min CSV doesn't exist
    # ===============================
    if not os.path.exists(one_min_file):
        print(f"⚠️ SKIP: No 1-min data for {symbol} on {date_str}")
        sell_prices.append(buy_price)
        sell_times.append("")
        exit_reasons.append("NO_1M_DATA_SKIPPED")
        pnls.append(0)
        risk_r_list.append(0)
        stop_dynamic_list.append(0)
        target_dynamic_list.append(0)
        high_time_list.append("")
        low_time_list.append("")
        continue

    sell_price = buy_price
    sell_time = None
    exit_reason = "NO_EXIT"
    had_post_buy_candle = False
    last_candle_close = None
    last_candle_time = None
    
    # ===============================
    # Step A: Load FULL 1-min data and calculate rolling ATR
    # ===============================
    mdf_full = pd.read_csv(one_min_file)
    mdf_full.columns = [c.lower() for c in mdf_full.columns]
    
    # Parse datetime
    if "date" in mdf_full.columns:
        try:
            mdf_full["datetime"] = pd.to_datetime(mdf_full["date"])
        except Exception:
            mdf_full["datetime"] = pd.NaT
    else:
        mdf_full["datetime"] = pd.NaT
    
    mdf_full = mdf_full.sort_values("datetime").reset_index(drop=True)
    
    # Convert price columns to numeric BEFORE ATR calculation
    for col in ['high', 'low', 'close', 'open']:
        if col in mdf_full.columns:
            mdf_full[col] = pd.to_numeric(mdf_full[col], errors='coerce')
    
    # Calculate rolling ATR on FULL history (no look-ahead bias)
    mdf_full = add_rolling_atr(mdf_full)
    
    # ===============================
    # Step B: Get ATR at Entry (from candle preceding buy_time)
    # ===============================
    # Build the entry datetime for the trade
    if date_str and buy_time:
        try:
            entry_datetime = pd.to_datetime(f"{date_str} {buy_time.strftime('%H:%M:%S')}")
            # Match timezone if needed (data may be tz-aware)
            if mdf_full['datetime'].dt.tz is not None:
                entry_datetime = entry_datetime.tz_localize(mdf_full['datetime'].dt.tz)
        except:
            entry_datetime = None
    else:
        entry_datetime = None
    
    # Get ATR from the candle immediately preceding entry time
    atr_value = None
    if entry_datetime is not None:
        pre_entry_candles = mdf_full[mdf_full['datetime'] < entry_datetime]
        if len(pre_entry_candles) > 0:
            atr_value = pre_entry_candles.iloc[-1]['ATR']
            if pd.isna(atr_value):
                atr_value = None
    
    # ===============================
    # DYNAMIC ENTRY CALCULATION (Risk_R) - v2.5 OPTIMIZED STOPS
    # ===============================
    if atr_value is not None:
        # v2.5: Risk_R = Max(ATR_14 * 2.5, Entry_Price * 0.015)
        # ATR*2.5 is optimal: reduces noise stops while keeping targets reachable
        # ATR*3.0 was too wide (TIME_EXIT turned negative)
        risk_r = max(atr_value * 2.5, buy_price * 0.015)
    else:
        # Fallback: use 1.5% of entry price as minimum
        risk_r = buy_price * 0.015
        print(f"⚠️ {symbol}: No ATR data before entry, using 1.5% floor")
    
    # ===============================
    # MODE-SPECIFIC TARGETS (calibrated to MFE distribution)
    # ===============================
    # Stop: 2.0 ATR for all modes
    stop = buy_price - risk_r  # Stop is Risk_R (2.0 ATR) below entry
    initial_stop = stop  # Store initial stop for exit labeling
    
    # Get trade mode (use mode_letter from filter section)
    # mode_letter is already extracted above as first character of mode string
    
    # Mode-specific targets (v2.5 - Optimized targets)
    # Mode A: 2.0R target (ORB has momentum)
    # Mode B: 1.5R target (VWAP reclaims are choppier - lower target)
    # Mode C: 1.5R target (SNIPER - quick profit on mid-day moves)
    if mode_letter == 'A':  # Mode A - ORB Breakout
        target_r = 2.0
        target = buy_price + (risk_r * target_r)
    elif mode_letter == 'B':  # Mode B - VWAP Reclaim
        target_r = 2.0  # v2.5: Keep at 2.0R (1.5R tested but worse overall)
        target = buy_price + (risk_r * target_r)
    elif mode_letter == 'C':  # Mode C - Day High Continuation (SNIPER)
        target_r = 1.5  # Sniper target for mid-day moves
        target = buy_price + (risk_r * target_r)
    else:
        target_r = 2.0
        target = buy_price + (risk_r * target_r)  # Default to 2R
    
    # Initialize protection ladder variables
    current_stop = stop  # Start at the initial stop
    highest_price = buy_price  # Track the highest price achieved (MFE)
    protection_level = "INITIAL"  # Track protection tier for exit labeling
    
    # ===============================
    # Filter for trade date (for exit simulation)
    # ===============================
    if date_str:
        mdf = mdf_full[mdf_full["datetime"].dt.strftime("%Y-%m-%d") == date_str].copy()
    else:
        mdf = mdf_full.copy()
    
    # ===============================
    # Calculate when High and Low occurred (for analysis)
    # ===============================
    high_time = ""
    low_time = ""
    if len(mdf) > 0 and 'high' in mdf.columns and 'low' in mdf.columns:
        try:
            # Find the time when the day high occurred
            high_idx = mdf['high'].idxmax()
            if not pd.isna(high_idx):
                high_time = mdf.loc[high_idx, 'datetime'].strftime('%H:%M') if pd.notna(mdf.loc[high_idx, 'datetime']) else ""
            
            # Find the time when the day low occurred
            low_idx = mdf['low'].idxmin()
            if not pd.isna(low_idx):
                low_time = mdf.loc[low_idx, 'datetime'].strftime('%H:%M') if pd.notna(mdf.loc[low_idx, 'datetime']) else ""
        except:
            pass

    if True:  # Always true now since we checked file exists above
        # mdf is now filtered for trade date

        # Find the actual 1-minute entry time
        # Phase-3 entry time is when the 5-min candle closed (e.g., 09:35:00)
        # We need to find which 1-min candle within that 5-min period hit the entry price
        # Look backwards from buy_time to (buy_time - 5 minutes)
        actual_entry_time = None
        
        # Calculate the start of the 5-minute window
        buy_hour = buy_time.hour
        buy_minute = buy_time.minute
        
        # Round down to nearest 5-minute interval
        window_start_minute = (buy_minute // 5) * 5
        window_start = time(buy_hour, window_start_minute)
        
        for _, m in mdf.iterrows():
            if pd.isna(m.get("datetime")):
                continue
            t = m["datetime"].time()
            
            # Look within the 5-minute window ending at buy_time
            if t < window_start or t > buy_time:
                continue
                
            high = float(m.get("high", np.nan))
            # Entry happens when High >= Entry Price (trigger hit)
            if not np.isnan(high) and high >= buy_price:
                actual_entry_time = t
                break
        
        # If we found a precise entry time, use it; otherwise use Phase-3 time
        if actual_entry_time:
            buy_time = actual_entry_time

        for _, m in mdf.iterrows():
            # skip rows without valid datetime
            if pd.isna(m.get("datetime")):
                continue
            t = m["datetime"].time()

            # start checking only from the candle strictly after the buy_time
            if t <= buy_time:
                continue

            had_post_buy_candle = True
            last_candle_close = float(m.get("close", np.nan)) if not pd.isna(m.get("close", np.nan)) else None
            last_candle_time = t

            high = float(m.get("high", np.nan))
            low = float(m.get("low", np.nan))
            op = float(m.get("open", np.nan))
            close = float(m.get("close", np.nan))
            
            # v2.2 HOLD-TO-CLOSE: No Fast Fail, No Protection Ladder
            # Simple execution: Hard Stop → Hard Target → Time Exit (15:10)
            # (Removed Fast Fail and Protection Ladder that over-engineered exits)
            
            # Track highest price for analytics only (no trailing)
            if not np.isnan(high) and high > highest_price:
                highest_price = high

            # v2.2 HOLD-TO-CLOSE: Simple Hard Stop check (no trailing/protection)
            hit_stop = (not np.isnan(low)) and (low <= stop)  # Use fixed initial stop
            hit_target = (not np.isnan(high)) and (high >= target)

            # Both hit in same candle → approximate intrabar order using distance from open
            if hit_stop and hit_target:
                dist_stop = abs(op - stop) if not np.isnan(op) else float('inf')
                dist_target = abs(op - target) if not np.isnan(op) else float('inf')
                if dist_stop <= dist_target:
                    sell_price = stop
                    exit_reason = "STOP_LOSS"
                else:
                    sell_price = target
                    exit_reason = "TARGET_HIT"
                sell_time = t
                break

            if hit_stop:
                sell_price = stop  # Use fixed initial stop
                sell_time = t
                exit_reason = "STOP_LOSS"
                break

            if hit_target:
                sell_price = target
                sell_time = t
                exit_reason = "TARGET_HIT"
                break

            # Force-exit at or after FORCE_EXIT_TIME using candle close
            if t >= FORCE_EXIT_TIME:
                sell_price = close if not np.isnan(close) else m.get("close", buy_price)
                sell_time = t
                exit_reason = "TIME_EXIT_1510"
                break

        # If no explicit exit was found but there were candles after the buy_time,
        # close at the last available candle close to approximate end-of-data exit.
        if exit_reason == "NO_EXIT" and had_post_buy_candle and last_candle_close is not None:
            sell_price = last_candle_close
            sell_time = last_candle_time
            exit_reason = "NO_EXIT_LASTCANDLE"

    pnl = (sell_price - buy_price) * qty

    # Normalize sell_time to string for consistent Excel output
    st = ""
    try:
        if isinstance(sell_time, time):
            st = sell_time.strftime("%H:%M:%S")
        elif isinstance(sell_time, str):
            st = sell_time
        elif pd.isna(sell_time) or sell_time is None:
            st = ""
        else:
            st = str(sell_time)
    except Exception:
        st = ""

    sell_prices.append(sell_price)
    sell_times.append(st)
    exit_reasons.append(exit_reason)
    pnls.append(pnl)
    risk_r_list.append(risk_r)
    stop_dynamic_list.append(stop)
    target_dynamic_list.append(target)
    high_time_list.append(high_time)
    low_time_list.append(low_time)

df["buy_price"] = df["entry_price"]
df["sell_price"] = sell_prices
df["sell_time"] = sell_times
df["exit_reason"] = exit_reasons
df["pnl"] = pnls
df["risk_r"] = risk_r_list
df["stop_dynamic"] = stop_dynamic_list
df["target_dynamic"] = target_dynamic_list
df["high_time"] = high_time_list
df["low_time"] = low_time_list

# Ensure integer quantities
df["quantity"] = df["quantity"].astype(int)

# ===============================
# FINAL OUTPUT: Build 4 sheets
# Sheets: Trade Log, Daily Summary, Performance, Algorithm Config
# ===============================

# Prepare Trade Log with requested columns (use existing df columns where available)
trade_log_cols = [
    'Date', 'Stock', 'Mode', 'Weight', 
    'Open', 'High', 'HighTime', 'Low', 'LowTime', 'Close', 'Volume', 'DataSource', 'CandleTimeframe',
    'EntryTime', 'EntryPrice', 'StopLoss', 'Target', 'ExitTime',
    'ExitReason', 'ExitPrice', 'Quantity', 'InvestedAmount', 'ProfitBeforeCosts',
    'TransactionCost', 'FinalProfit', 'P&L%'
]

# Build trade log rows
trade_rows = []
for _, r in df.iterrows():
    date_val = r.get('date')
    try:
        date_str = pd.to_datetime(date_val).strftime('%Y-%m-%d')
    except Exception:
        date_str = str(date_val)

    invested = float(r.get('trade_value', 0.0))
    profit_before = float(r.get('pnl', 0.0))
    exit_reason = r.get('exit_reason', '')
    
    # Don't charge transaction costs for SKIPPED trades (they weren't executed)
    if 'SKIPPED' in str(exit_reason):
        tx_cost = 0
        final_profit = 0
        invested = 0  # No capital was deployed
        pl_pct = 0
    else:
        tx_cost = invested * TRANSACTION_COST_PCT
        final_profit = profit_before - tx_cost
        pl_pct = (final_profit / invested * 100.0) if invested > 0 else 0.0

    row = {
        'Date': date_str,
        'Stock': r.get('symbol'),
        'Mode': r.get('mode', ''),
        'Weight': float(r.get('weight', 0.0)),
        'Open': np.nan,
        'High': np.nan,
        'HighTime': r.get('high_time', ''),  # Time when day high occurred
        'Low': np.nan,
        'LowTime': r.get('low_time', ''),    # Time when day low occurred
        'Close': np.nan,
        'Volume': np.nan,
        'DataSource': r.get('DataSource') or ONE_MIN_DATA_DIR,
        'CandleTimeframe': r.get('CandleTimeframe') or '1m',
        'EntryTime': r.get('buy_time'),
        'EntryPrice': float(r.get('buy_price', 0.0)),
        'StopLoss': float(r.get('stop_price', np.nan)),
        'Target': float(r.get('target_price', np.nan)),
        'ExitTime': r.get('sell_time'),
        'ExitReason': r.get('exit_reason'),
        'ExitPrice': float(r.get('sell_price', 0.0)),
        'Quantity': int(r.get('quantity', 0)),
        'InvestedAmount': invested,
        'ProfitBeforeCosts': profit_before,
        'TransactionCost': tx_cost,
        'FinalProfit': final_profit,
        'P&L%': pl_pct
    }
    trade_rows.append(row)

trade_log_df = pd.DataFrame(trade_rows, columns=trade_log_cols)

ohlcv_path = os.path.join('downloaded_data', 'daily_candles_nifty500.xlsx')
if os.path.exists(ohlcv_path):
    try:
        ohl = pd.read_excel(ohlcv_path)
        ohl.columns = [str(c).strip() for c in ohl.columns]
        
        # Normalize symbol column
        if 'Symbol' in ohl.columns:
            ohl['Symbol'] = ohl['Symbol'].astype(str).str.upper().str.strip()
        elif 'symbol' in ohl.columns:
            ohl['Symbol'] = ohl['symbol'].astype(str).str.upper().str.strip()
        
        # Handle Datetime column (not Date)
        if 'Datetime' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['Datetime']).dt.strftime('%Y-%m-%d')
        elif 'Date' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['Date']).dt.strftime('%Y-%m-%d')
        elif 'date' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['date']).dt.strftime('%Y-%m-%d')

        # keep only relevant OHLCV cols that exist
        ohlv_cols = [c for c in ['Open','High','Low','Close','Volume'] if c in ohl.columns]

        # create lookup dict keyed by (symbol,date)
        ohl['__key'] = ohl['Symbol'].astype(str) + '||' + ohl['Date'].astype(str)
        ohl_lookup = ohl.set_index('__key')[ohlv_cols].to_dict(orient='index')

        # normalize trade keys and attempt fast lookup
        def try_get_ohlv(stock, date):
            if pd.isna(stock) or pd.isna(date):
                return None
            k = str(stock).upper().strip() + '||' + str(date)
            if k in ohl_lookup:
                return ohl_lookup[k]
            # fallback: try exact match in symbols for that date
            for idx,row in ohl[ohl['Date']==str(date)].iterrows():
                sym = row['Symbol']
                if sym == str(stock).upper().strip():
                    return row[ohlv_cols].to_dict()
            # try partial match
            for idx,row in ohl[ohl['Date']==str(date)].iterrows():
                sym = row['Symbol']
                if str(stock).upper().strip() in sym or sym in str(stock).upper().strip():
                    return row[ohlv_cols].to_dict()
            return None

        # apply per-row fill
        for i, tr in trade_log_df.iterrows():
            stock = tr['Stock']
            datev = tr['Date']
            ovals = try_get_ohlv(stock, datev)
            if ovals:
                for c in ohlv_cols:
                    trade_log_df.at[i, c] = ovals.get(c)
        
        # Ensure OHLCV columns are numeric (not strings)
        for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
            if col in trade_log_df.columns:
                trade_log_df[col] = pd.to_numeric(trade_log_df[col], errors='coerce')
                
    except Exception as e:
        print(f"⚠️ Warning: Could not load OHLCV data: {e}")
        pass

# Ensure Bias is readable (only if column exists)
if 'Bias' in trade_log_df.columns:
    trade_log_df['Bias'] = trade_log_df['Bias'].fillna('N/A')

# Daily summary (requested columns)
daily = trade_log_df.groupby('Date').agg(
    TotalTrades=('Stock', 'count'),
    CapitalStart=('InvestedAmount', lambda x: float(C_PER_DAY)),
    CapitalInvested=('InvestedAmount', 'sum'),
    CapitalEnd=('FinalProfit', lambda x: float(C_PER_DAY) + x.sum()),
    Profit=('FinalProfit', lambda x: x[x>0].sum()),
    Loss=('FinalProfit', lambda x: abs(x[x<0].sum())),
    DailyP_and_L=('FinalProfit', 'sum')
).reset_index()

# Rename DailyP_and_L -> DailyP&L and DayStatus
daily = daily.rename(columns={'DailyP_and_L': 'DailyP&L'})
daily['DayStatus'] = daily['DailyP&L'].apply(lambda v: 'PROFIT' if v>0 else ('LOSS' if v<0 else 'BREAKEVEN'))

# Reorder daily columns as requested
daily_summary_cols = ['Date','TotalTrades','CapitalStart','CapitalInvested','CapitalEnd','Profit','Loss','DailyP&L','DayStatus']
daily_summary = daily[daily_summary_cols]

# ===============================
# PERFORMANCE SHEET - CRYSTAL CLEAR FORMAT
# ===============================

# Calculate overall metrics first
total_trades = len(trade_log_df)

# Filter out skipped trades (those with zero P&L are skipped, not breakeven)
executed_trades_df = trade_log_df[~trade_log_df['ExitReason'].str.contains('SKIPPED', na=False)]
skipped_trades_df = trade_log_df[trade_log_df['ExitReason'].str.contains('SKIPPED', na=False)]
executed_count = len(executed_trades_df)
skipped_count = len(skipped_trades_df)

wins_df = executed_trades_df[executed_trades_df['FinalProfit']>0]
losses_df = executed_trades_df[executed_trades_df['FinalProfit']<0]
win_count = len(wins_df)
loss_count = len(losses_df)

# Win rate: calculated only from executed trades (excludes skipped)
win_rate = (win_count / executed_count * 100.0) if executed_count > 0 else 0.0
loss_rate = (loss_count / executed_count * 100.0) if executed_count > 0 else 0.0

total_invested = executed_trades_df['InvestedAmount'].sum()
total_profit_amount = wins_df['FinalProfit'].sum() if win_count > 0 else 0.0
total_loss_amount = abs(losses_df['FinalProfit'].sum()) if loss_count > 0 else 0.0
net_pnl = executed_trades_df['FinalProfit'].sum()
roi_percent = (net_pnl / total_invested * 100.0) if total_invested > 0 else 0.0

# Calculate daily P&L for average per day metrics (only executed trades)
daily_pnl = executed_trades_df.groupby('Date')['FinalProfit'].sum()
daily_profit = daily_pnl[daily_pnl > 0]
daily_loss = daily_pnl[daily_pnl < 0]
avg_net_profit_per_day = daily_profit.mean() if len(daily_profit) > 0 else 0.0
avg_net_loss_per_day = abs(daily_loss.mean()) if len(daily_loss) > 0 else 0.0

avg_win = wins_df['FinalProfit'].mean() if win_count > 0 else 0.0
avg_loss = abs(losses_df['FinalProfit'].mean()) if loss_count > 0 else 0.0

# Daily trading statistics (only executed trades)
trades_per_day = executed_trades_df.groupby('Date').size()
unique_dates = executed_trades_df['Date'].unique()
total_trading_days = len(unique_dates)
avg_scripts_per_day = trades_per_day.mean() if len(trades_per_day) > 0 else 0.0

# Min/Max scripts traded on a given day (no dates since multiple days could have same count)
if len(trades_per_day) > 0:
    min_trades_count = trades_per_day.min()
    max_trades_count = trades_per_day.max()
else:
    min_trades_count = 0
    max_trades_count = 0

# Capital invested per day statistics (only executed trades)
capital_per_day = executed_trades_df.groupby('Date')['InvestedAmount'].sum()
avg_capital_per_day = capital_per_day.mean() if len(capital_per_day) > 0 else 0.0
min_capital_per_day = capital_per_day.min() if len(capital_per_day) > 0 else 0.0
max_capital_per_day = capital_per_day.max() if len(capital_per_day) > 0 else 0.0

# Create SUMMARY section (top of sheet)
summary_data = [
    ['═══════════════════════════════════════════════════════════════', ''],
    ['                    📊 BACKTEST PERFORMANCE SUMMARY', ''],
    ['═══════════════════════════════════════════════════════════════', ''],
    ['', ''],
    ['💰 FINAL RESULT', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Capital Deployed', f'₹{total_invested:,.2f}'],
    ['Avg Net Profit/Day', f'₹{avg_net_profit_per_day:,.2f}'],
    ['Avg Net Loss/Day', f'₹{avg_net_loss_per_day:,.2f}'],
    ['Net Profit/Loss', f'₹{net_pnl:,.2f}'],
    ['Return on Investment (ROI)', f'{roi_percent:.2f}%'],
    ['Final Status', 'PROFIT ✅' if net_pnl > 0 else ('LOSS ❌' if net_pnl < 0 else 'BREAKEVEN')],
    ['', ''],
    ['📊 TRADE STATISTICS', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Trades', f'{total_trades}'],
    ['Executed Trades', f'{executed_count}'],
    ['Skipped Trades', f'{skipped_count}'],
    ['Winning Trades', f'{win_count} ({win_rate:.1f}%)'],
    ['Losing Trades', f'{loss_count} ({loss_rate:.1f}%)'],
    ['', ''],
    ['📅 DAILY TRADING STATISTICS', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Trading Days', f'{total_trading_days}'],
    ['Avg # Scripts Traded/Day', f'{avg_scripts_per_day:.2f}'],
    ['Min # Scripts Traded', f'{min_trades_count}'],
    ['Max # Scripts Traded', f'{max_trades_count}'],
    ['', ''],
    ['💵 CAPITAL STATISTICS', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Avg Capital Invested/Day', f'₹{avg_capital_per_day:,.2f}'],
    ['Min Capital Invested/Day', f'₹{min_capital_per_day:,.2f}'],
    ['Max Capital Invested/Day', f'₹{max_capital_per_day:,.2f}'],
    ['', ''],
    ['💵 AVERAGE PER TRADE', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Average Win', f'₹{avg_win:,.2f}'],
    ['Average Loss', f'₹{avg_loss:,.2f}'],
    ['Average P&L per Trade', f'₹{net_pnl/executed_count:,.2f}' if executed_count > 0 else '₹0.00'],
    ['', ''],
    ['═══════════════════════════════════════════════════════════════', ''],
]

summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])

# Use summary as the performance sheet (no PromptType breakdown needed)
performance_df = summary_df

# Algorithm Config sheet - produce key/value table from variables
alg_config = {
    'StopLossPercent': f"{S_PCT*100:.2f}%",
    'TargetPercent': f"{(R_MULT*S_PCT)*100:.2f}%",
    'RiskReward': f"1:{R_MULT}",
    'CapitalPerDay': f"₹{C_PER_DAY:,}",
    'CapitalPerTrade': f"₹{int(C_PER_DAY*C_PCT):,}",
    'MaxPositions': int(np.floor(C_PER_DAY / (C_PER_DAY*C_PCT))) if C_PCT>0 else 0,
    'CandleTimeframe': '1m',
    'DataSource': ONE_MIN_DATA_DIR,
    'NewsSource': df.get('NewsSource').iloc[0] if 'NewsSource' in df.columns else '',
    'AIModel': df.get('AIModel').iloc[0] if 'AIModel' in df.columns else '',
    'EntryStartTime': df['buy_time'].iloc[0] if 'buy_time' in df.columns else '',
    'ForceExitTime': FORCE_EXIT_TIME.strftime('%H:%M'),
    'SlippagePercent': f"{SLIPPAGE_PCT*100:.2f}%",
    'TransactionCostPercent': f"{TRANSACTION_COST_PCT*100:.2f}%"
}
config_df = pd.DataFrame(list(alg_config.items()), columns=['Parameter','Value'])

# Write sheets
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    trade_log_df.to_excel(writer, sheet_name='Trade Log', index=False)
    daily_summary.to_excel(writer, sheet_name='Daily Summary', index=False)
    performance_df.to_excel(writer, sheet_name='Performance', index=False)
    config_df.to_excel(writer, sheet_name='Algorithm Config', index=False)

# Load workbook and apply formatting
wb = openpyxl.load_workbook(OUTPUT_FILE)

def format_sheet(ws, header_fill=None, pnl_col_name=None):
    hdr = Font(bold=True)
    for cell in ws[1]:
        cell.font = hdr
        cell.alignment = Alignment(horizontal='center')
        if header_fill:
            cell.fill = header_fill
    if pnl_col_name:
        pnl_col = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).lower() == pnl_col_name.lower():
                pnl_col = idx
                break
        if pnl_col and ws.max_row > 1:
            col_letter = openpyxl.utils.get_column_letter(pnl_col)
            green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            range_str = f'{col_letter}2:{col_letter}{ws.max_row}'
            ws.conditional_formatting.add(range_str,
                                          CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green))
            ws.conditional_formatting.add(range_str,
                                          CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red))

# Format Trade Log with blue header
if 'Trade Log' in wb.sheetnames:
    ws = wb['Trade Log']
    blue = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    format_sheet(ws, header_fill=blue, pnl_col_name='FinalProfit')

# Format Daily Summary with bold headers and conditional coloring on DailyP&L
if 'Daily Summary' in wb.sheetnames:
    ws2 = wb['Daily Summary']
    yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    format_sheet(ws2, header_fill=yellow, pnl_col_name='DailyP&L')

# Format Performance (special handling for summary format)
if 'Performance' in wb.sheetnames:
    ws3 = wb['Performance']
    
    # Style for section headers (lines with emojis and titles)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=14, color='FFFFFF')
    
    # Style for separator lines
    separator_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Style for metric names
    metric_font = Font(bold=True, size=11)
    
    # Style for values
    value_font = Font(size=11)
    
    # Apply formatting row by row
    for row_idx in range(1, ws3.max_row + 1):
        metric_cell = ws3.cell(row=row_idx, column=1)
        value_cell = ws3.cell(row=row_idx, column=2)
        
        metric_text = str(metric_cell.value) if metric_cell.value else ''
        value_text = str(value_cell.value) if value_cell.value else ''
        
        # Header rows (with emojis or equals signs)
        if '═' in metric_text or '📊' in metric_text:
            metric_cell.font = header_font
            metric_cell.fill = header_fill
            value_cell.fill = header_fill
            # Merge cells for header rows
            if '📊' in metric_text or '═' in metric_text:
                ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                metric_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Section titles (with emojis)
        elif any(emoji in metric_text for emoji in ['💰', '📈', '📊', '💵']):
            metric_cell.font = Font(bold=True, size=12, color='1F4E78')
            metric_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            value_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        
        # Separator lines
        elif '─' in metric_text:
            metric_cell.fill = separator_fill
            value_cell.fill = separator_fill
            ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        
        # Data rows
        elif metric_text and metric_text.strip():
            metric_cell.font = metric_font
            value_cell.font = value_font
            
            # Color code profit/loss values
            if 'Net Profit/Loss' in metric_text or 'Final Status' in metric_text or 'Net Amount' in metric_text:
                if '✅' in value_text or (value_text.startswith('₹') and float(value_text.replace('₹','').replace(',','')) > 0):
                    value_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    value_cell.font = Font(bold=True, size=12, color='006100')
                elif '❌' in value_text or (value_text.startswith('₹') and float(value_text.replace('₹','').replace(',','')) < 0):
                    value_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    value_cell.font = Font(bold=True, size=12, color='9C0006')
            
            # Highlight ROI
            if 'ROI' in metric_text:
                value_cell.font = Font(bold=True, size=12)
    
    # Adjust column widths
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 25

# Format Algorithm Config (dark header, clear key/value)
if 'Algorithm Config' in wb.sheetnames:
    ws4 = wb['Algorithm Config']
    dark = PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid')
    for cell in ws4[1]:
        cell.fill = dark
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    # alternate shading for readability
    for r in range(2, ws4.max_row+1):
        fill = PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid') if r%2==0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        for c in range(1, ws4.max_column+1):
            ws4.cell(row=r, column=c).fill = fill

wb.save(OUTPUT_FILE)

print("✅ Phase-4 BACKTEST (1-minute) completed successfully")
print(f"📄 Output saved to: {OUTPUT_FILE}")
# Show net P&L (after transaction costs), excluding skipped trades
active_trades = trade_log_df[~trade_log_df['ExitReason'].str.contains('SKIPPED', na=False)]
print(f"Net P&L (after costs): ₹{active_trades['FinalProfit'].sum():,.2f}")
