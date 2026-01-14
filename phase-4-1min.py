"""
Project: Pulse915 Backtesting System
Phase: 4 (1-Min) - High-Precision Backtesting & Execution Simulation
Version: 3.5
Date: 2026-01-12
Updated By: Sabrish Surender

Description:
Architecture: "Regime-Gated Execution with Consistent Geometry & Progress Logic"

THE THREE PILLARS:
1. Consistent Geometry: Phase 4 is a "Slave" to Phase 3. We use Entry/Stop/Target 
   exactly as defined. No artificial floors, no ATR recalculation.

2. Regime Awareness: Alpha is not evenly distributed. We mechanically block trades
   during low-quality time zones (Midday Chop, Late Fades).

3. Progress Physics: A trade is an entity with momentum. If it fails to achieve
   X velocity (R-multiple) within T time, it's dead. Exit on stagnation.

Regime Rules:
- Mode A (ORB): 09:15-10:15 ONLY (morning momentum edge decays)
- Mode B (VWAP): Bookends only: 09:15-11:00 + 13:30-15:00 (kill 11:00-13:30)
- Mode C (Structure): 10:00-14:00 (need time for day high, stop before low volume)

Exit Logic:
- Half-Risk Ladder: At +1.0R MFE → Stop = Entry - 0.5R, At +2.0R → Stop = Entry + 0.5R
- No-Progress Engine: Mode-specific time limits for stagnation detection
- Kill Switch: Daily loss limit of ₹10,000 (1% circuit breaker)

Capital Sizing: Fixed ₹5,00,000 per trade (prevents leverage blowups)
"""


# phase-4.py
import os
import sys
import pandas as pd
import numpy as np
import warnings
from datetime import time, datetime

# Suppress pandas datetime parsing warnings
warnings.filterwarnings('ignore', message='Could not infer format')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
# ===============================
# CONFIG
# ===============================
PHASE3_FILE = "phase-3results/Phase3_results.xlsx"
ONE_MIN_DATA_DIR = "downloaded_data/1min"   # 1-minute candles
OUTPUT_DIR = "phase-4results"
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"phase4_results_1m_{ts}.xlsx")

# ===============================
# v3.5 CAPITAL CONFIG
# ===============================
C_PER_DAY = 1000000       # ₹10,00,000 per day total
MAX_NOTIONAL = 500000     # ₹5,00,000 max notional per trade (cap)
RISK_PER_TRADE = 2500     # ₹2,500 max risk per trade (controls position size)
FORCE_EXIT_TIME = time(15, 10)

# ===============================
# v3.5 MODULE 2: REGIME GATEKEEPER
# ===============================
# Mode A (ORB): 09:15-10:15 ONLY
# Mode B (VWAP): Bookends: 09:15-11:00 + 13:30-15:00 (kill 11:00-13:30)
# Mode C (Structure): 10:00-14:00
REGIME_WINDOWS = {
    'A': {'allowed': [(time(9, 15), time(10, 15))]},
    'B': {'allowed': [(time(9, 15), time(11, 0)), (time(13, 30), time(15, 0))]},
    'C': {'allowed': [(time(10, 0), time(14, 0))]},
}

# ===============================
# v3.5 MODULE 3: NO-PROGRESS ENGINE (DISABLED for now)
# ===============================
# Theory: Breakouts have a "half-life." No progress = exit.
NO_PROGRESS_CONFIG = {
    'A': {'time_limit': 25, 'threshold': 0.30},  # Must show impulse fast
    'B': {'time_limit': 30, 'threshold': 0.25},  # Slower grind allowed
    'C': {'time_limit': 18, 'threshold': 0.35},  # Breakouts must work fast
}

# ===============================
# v3.5 MODULE 5: PORTFOLIO KILL SWITCH
# ===============================
DAILY_LOSS_LIMIT = -10000  # ₹10,000 max daily loss (1% circuit breaker)

# ===============================
# v3.5 MODULE 6: REALISTIC FRICTION
# ===============================
TRANSACTION_COST_PCT = 0.0005  # 0.05% of turnover
SLIPPAGE_PCT = 0.001           # 0.1% applied to BOTH entry and exit

# Minimum edge filter (prevents "death by a thousand paper cuts")
COST_BUFFER = 0.0025  # 0.25% minimum profit after costs

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ===============================
# ATR CALCULATION FOR DYNAMIC RISK
# ===============================
def add_rolling_atr(df_full):
    """
    Calculate rolling 14-period ATR on the entire historical dataset.
    Uses Wilder's smoothing method (EMA with alpha=1/14).
    Adds 'ATR' column to the dataframe.
    
    Args:
        df_full: DataFrame with 'high', 'low', 'close' columns (lowercase)
    
    Returns:
        DataFrame with 'ATR' column added
    """
    df = df_full.copy()
    df['prev_close'] = df['close'].shift(1)
    df['tr'] = np.maximum(
        df['high'] - df['low'],
        np.maximum(
            abs(df['high'] - df['prev_close']),
            abs(df['low'] - df['prev_close'])
        )
    )
    # Wilder's EMA (span=14)
    df['ATR'] = df['tr'].ewm(span=14, adjust=False).mean()
    
    # Clean up temporary columns
    df = df.drop(columns=['prev_close', 'tr'], errors='ignore')
    return df

# ===============================
# v5.0 MODULE 2: REGIME GATEKEEPER
# ===============================
def check_regime_allowed(mode_letter, fill_time):
    """
    Module 2: Regime Gatekeeper
    Alpha is not distributed evenly. Block trades during "Kill Zones."
    
    Returns: (allowed, reason)
    """
    if mode_letter not in REGIME_WINDOWS:
        return True, None  # Unknown mode, allow by default
    
    windows = REGIME_WINDOWS[mode_letter]['allowed']
    
    for start, end in windows:
        if start <= fill_time <= end:
            return True, None
    
    # Not in any allowed window
    return False, f"REGIME_FILTER_{mode_letter}"

# ===============================
# v5.0 MODULE 3: NO-PROGRESS ENGINE
# ===============================
def check_no_progress(minutes_in_trade, mfe_r, mode_letter):
    """
    Module 3: No-Progress Engine
    Theory: A breakout trade has a "half-life." If it doesn't move significantly
    within T minutes, the probability of failure approaches 80%.
    
    Returns: (should_exit, reason)
    """
    config = NO_PROGRESS_CONFIG.get(mode_letter, NO_PROGRESS_CONFIG['A'])
    time_limit = config['time_limit']
    threshold = config['threshold']
    
    if minutes_in_trade >= time_limit and mfe_r < threshold:
        return True, f"NO_PROGRESS_{int(minutes_in_trade)}m_MFE{mfe_r:.2f}R"
    return False, None

# ===============================
# v5.0 MODULE 4: HALF-RISK LADDER
# ===============================
def calculate_half_risk_ladder(entry_price, risk_r, mfe_r, current_stop):
    """
    Module 4: The "Half-Risk" Ladder
    Step 1: At MFE >= 1.0R → Stop = Entry - 0.5R 
            (Survive the 1.12R noise, but halve the risk)
    Step 2: At MFE >= 2.0R → Stop = Entry + 0.5R 
            (Lock in "Intern's Salary")
    
    Returns: new_stop
    """
    new_stop = current_stop
    
    if mfe_r >= 2.0:
        # Step 2: Lock profit at +0.5R
        locked_stop = entry_price + (0.5 * risk_r)
        new_stop = max(new_stop, locked_stop)
    elif mfe_r >= 1.0:
        # Step 1: Compress risk to -0.5R
        compressed_stop = entry_price - (0.5 * risk_r)
        new_stop = max(new_stop, compressed_stop)
    
    return new_stop


# ===============================
# LOAD PHASE-3 OUTPUT
# ===============================
df = pd.read_excel(PHASE3_FILE)

df = df.rename(columns={
    "Stock": "symbol",
    "Entry Price (₹)": "entry_price",
    "Stop-Loss (₹)": "stop_loss",  # From Phase-3
    "Target (₹)": "target",        # From Phase-3
    "Date": "date",
    "Entry Time": "buy_time",
    "Entry Mode": "mode"  # A, B, or C from Phase-3
})

# Read VelocityScore if available, otherwise use default
if "VelocityScore" in df.columns:
    df["velocity_score"] = pd.to_numeric(df["VelocityScore"], errors="coerce").fillna(75.0)
else:
    # Default VelocityScore if not provided by Phase-3
    # TODO: Phase-3 should output VelocityScore in future
    df["velocity_score"] = 75.0

required_cols = {"symbol", "entry_price", "stop_loss", "target", "date", "buy_time"}
# Prepare output: multi-sheet Excel with formatting
missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"❌ Missing required columns: {missing}")

# ===============================
# MINIMUM EDGE FILTER
# ===============================
# Filter: (Target - Entry) / Entry >= CostBuffer
# Prevents "death by a thousand paper cuts"
initial_count = len(df)
df["potential_profit_pct"] = ((df["target"] - df["entry_price"]) / df["entry_price"])
df = df[df["potential_profit_pct"] >= COST_BUFFER].copy()
filtered_count = initial_count - len(df)
print(f"Minimum edge filter: Removed {filtered_count} trades with <{COST_BUFFER*100:.2f}% potential profit")
print(f"Remaining trades: {len(df)}")
df = df.drop(columns=["potential_profit_pct"])  # Clean up temporary column

# ===============================
# v3.5 CONSISTENT GEOMETRY: USE PHASE-3 LEVELS EXACTLY
# ===============================
# Pillar 1: Phase 4 is a "Slave" to Phase 3
df["stop_price"] = pd.to_numeric(df["stop_loss"], errors="coerce")
df["target_price"] = pd.to_numeric(df["target"], errors="coerce")
df["entry_price"] = pd.to_numeric(df["entry_price"], errors="coerce")

# Calculate R (Risk Unit) for each trade
df["risk_r"] = df["entry_price"] - df["stop_price"]

# Validate geometry: R must be positive and meaningful
invalid_geometry = (df["risk_r"] <= 0) | (df["risk_r"] < df["entry_price"] * 0.001)
invalid_count = invalid_geometry.sum()
if invalid_count > 0:
    print(f"⚠️ v3.5 Geometry: Rejecting {invalid_count} trades with invalid R")
    df = df[~invalid_geometry].copy()

# ===============================
# v3.5 CAPITAL ALLOCATION: RISK-BASED SIZING + NOTIONAL CAP
# ===============================
# Fix C from ChatGPT: Fixed capital causes variable risk which violates daily limit
# Solution: Size by RISK (₹2,500 max loss at stop) + cap notional at ₹5L

# Friction buffer per share (for sizing safety)
friction_buffer = df["entry_price"] * (SLIPPAGE_PCT * 2 + TRANSACTION_COST_PCT)

# Qty from risk budget: qty = floor(RISK_PER_TRADE / (R + friction))
df["qty_risk"] = np.floor(RISK_PER_TRADE / (df["risk_r"] + friction_buffer))

# Qty from notional cap: qty = floor(MAX_NOTIONAL / entry)
df["qty_notional"] = np.floor(MAX_NOTIONAL / df["entry_price"])

# Take the minimum of the two (risk controls size, notional caps leverage)
df["quantity"] = np.minimum(df["qty_risk"], df["qty_notional"]).astype(int)
df["trade_value"] = df["quantity"] * df["entry_price"]

# Remove zero-quantity rows
df = df[df["quantity"] > 0].copy()

# For compatibility with output, add weight column
df["weight"] = 1.0 / len(df) if len(df) > 0 else 0.0

print(f"v3.5 Capital: Risk-based (₹{RISK_PER_TRADE:,}/trade risk), {len(df)} trades ready")
print(f"  Avg qty: {df['quantity'].mean():.0f}, Avg notional: ₹{df['trade_value'].mean():,.0f}")

# ===============================
# v3.5 MODULE 5: SORT BY FILL TIME FOR KILL SWITCH
# ===============================
df = df.sort_values(['date', 'buy_time']).reset_index(drop=True)



# ===============================
# v3.5 EXIT RESOLUTION (1-MIN BACKTEST)
# ===============================
# Daily P&L tracking for Kill Switch (Module 5)
daily_pnl_tracker = {}  # {date_str: cumulative_pnl}

sell_prices = []
sell_times = []
exit_reasons = []
pnls = []
txn_cost_list = []        # v3.5: Track ACTUAL transaction costs from engine
risk_r_list = []          # Dynamic risk values for verification
stop_dynamic_list = []    # Dynamic stop values for verification
target_dynamic_list = []  # Dynamic target values for verification
high_time_list = []       # Time when day high occurred
low_time_list = []        # Time when day low occurred
qty_executed_list = []    # v3.5: Track executed qty (0 for SKIPs)


for _, row in df.iterrows():
    symbol = row["symbol"]
    buy_price = row["entry_price"]
    stop = row["stop_price"]
    target = row["target_price"]
    qty = row["quantity"]
    buy_time = row["buy_time"]

    # normalize buy_time to a time object if possible
    if pd.isna(buy_time):
        buy_time = time(0, 0)
    elif isinstance(buy_time, str):
        try:
            bt = pd.to_datetime(buy_time).time()
            buy_time = bt
        except Exception:
            try:
                parts = [int(x) for x in buy_time.split(":")]
                buy_time = time(*parts[:3])
            except Exception:
                buy_time = time(0, 0)
    elif isinstance(buy_time, pd.Timestamp):
        buy_time = buy_time.time()

    # ===============================
    # EARLY CALCULATION: Get HighTime/LowTime for all trades (including skipped)
    # ===============================
    # This runs before filters so skipped trades also have this data
    date_val_early = row.get("date")
    date_str_early = None
    if not pd.isna(date_val_early):
        try:
            date_str_early = pd.to_datetime(date_val_early).strftime("%Y-%m-%d")
        except:
            date_str_early = str(date_val_early)
    
    one_min_file_early = os.path.join(ONE_MIN_DATA_DIR, f"{symbol}.csv")
    high_time = ""
    low_time = ""
    
    if os.path.exists(one_min_file_early):
        try:
            mdf_early = pd.read_csv(one_min_file_early)
            # Handle both 'datetime' and 'date' column names
            datetime_col = 'datetime' if 'datetime' in mdf_early.columns else 'date' if 'date' in mdf_early.columns else None
            if datetime_col:
                mdf_early[datetime_col] = pd.to_datetime(mdf_early[datetime_col], errors='coerce')
                if date_str_early:
                    mdf_early = mdf_early[mdf_early[datetime_col].dt.strftime('%Y-%m-%d') == date_str_early]
                
                if len(mdf_early) > 0 and 'high' in mdf_early.columns and 'low' in mdf_early.columns:
                    # Find when day high occurred
                    high_idx = mdf_early['high'].idxmax()
                    if not pd.isna(high_idx):
                        high_time = mdf_early.loc[high_idx, datetime_col].strftime('%H:%M') if pd.notna(mdf_early.loc[high_idx, datetime_col]) else ""
                    
                    # Find when day low occurred
                    low_idx = mdf_early['low'].idxmin()
                    if not pd.isna(low_idx):
                        low_time = mdf_early.loc[low_idx, datetime_col].strftime('%H:%M') if pd.notna(mdf_early.loc[low_idx, datetime_col]) else ""
        except Exception as e:
            pass  # Keep empty strings

    # ===============================
    # v5.0 MODULE 2: REGIME GATEKEEPER
    # ===============================
    # Get trade mode for filtering (extract first letter/word)
    trade_mode = row.get('mode', 'A')
    if pd.isna(trade_mode):
        trade_mode = 'A'
    trade_mode_str = str(trade_mode).strip()
    mode_letter = trade_mode_str[0].upper() if trade_mode_str else 'A'
    
    # Check if trade is in allowed regime window
    regime_allowed, regime_reason = check_regime_allowed(mode_letter, buy_time)
    if not regime_allowed:
        print(f"⚠️ SKIP: Mode {mode_letter} trade for {symbol} at {buy_time} ({regime_reason})")
        # v3.5 TRUE SKIP: No execution, no costs, no P&L
        sell_prices.append(np.nan)
        sell_times.append("")
        exit_reasons.append(regime_reason)
        pnls.append(0)
        txn_cost_list.append(0)  # No costs for skipped trades
        risk_r_list.append(row.get("risk_r", 0))
        stop_dynamic_list.append(stop)
        target_dynamic_list.append(target)
        high_time_list.append(high_time)
        low_time_list.append(low_time)
        qty_executed_list.append(0)  # No execution
        continue

    
    # ===============================
    # v5.0 MODULE 5: PORTFOLIO KILL SWITCH
    # ===============================
    # Build date string for daily tracking
    date_val = row.get("date")
    date_str = None
    if not pd.isna(date_val):
        try:
            date_str = pd.to_datetime(date_val).strftime("%Y-%m-%d")
        except Exception:
            date_str = str(date_val)
    
    if date_str:
        current_daily_pnl = daily_pnl_tracker.get(date_str, 0)
        if current_daily_pnl <= DAILY_LOSS_LIMIT:
            print(f"🛑 KILL SWITCH: {symbol} - Daily drawdown limit reached (₹{current_daily_pnl:,.0f})")
            # v3.5 TRUE SKIP: No execution, no costs, no P&L
            sell_prices.append(np.nan)
            sell_times.append("")
            exit_reasons.append("DAILY_DRAWDOWN_LIMIT")
            pnls.append(0)
            txn_cost_list.append(0)  # No costs for skipped trades
            risk_r_list.append(row.get("risk_r", 0))
            stop_dynamic_list.append(stop)
            target_dynamic_list.append(target)
            high_time_list.append(high_time)
            low_time_list.append(low_time)
            qty_executed_list.append(0)  # No execution
            continue



    # Build path to 1-minute CSV: downloaded_data/1min_candles/<SYMBOL>.csv
    date_val = row.get("date")
    date_str = None
    if pd.isna(date_val):
        date_str = None
    else:
        try:
            date_ts = pd.to_datetime(date_val)
            date_str = date_ts.strftime("%Y-%m-%d")
        except Exception:
            date_str = str(date_val)

    one_min_file = os.path.join(ONE_MIN_DATA_DIR, f"{symbol}.csv")

    # ===============================
    # SAFETY CHECK: Skip if 1-min CSV doesn't exist
    # ===============================
    if not os.path.exists(one_min_file):
        print(f"⚠️ SKIP: No 1-min data for {symbol} on {date_str}")
        # v3.5 TRUE SKIP: No execution, no costs, no P&L
        sell_prices.append(np.nan)
        sell_times.append("")
        exit_reasons.append("NO_1M_DATA_SKIPPED")
        pnls.append(0)
        txn_cost_list.append(0)
        risk_r_list.append(0)
        stop_dynamic_list.append(0)
        target_dynamic_list.append(0)
        high_time_list.append("")
        low_time_list.append("")
        qty_executed_list.append(0)
        continue


    sell_price = buy_price
    sell_time = None
    exit_reason = "NO_EXIT"
    had_post_buy_candle = False
    last_candle_close = None
    last_candle_time = None
    
    # ===============================
    # Step A: Load FULL 1-min data and calculate rolling ATR
    # ===============================
    mdf_full = pd.read_csv(one_min_file)
    mdf_full.columns = [c.lower() for c in mdf_full.columns]
    
    # Parse datetime
    if "date" in mdf_full.columns:
        try:
            mdf_full["datetime"] = pd.to_datetime(mdf_full["date"])
        except Exception:
            mdf_full["datetime"] = pd.NaT
    else:
        mdf_full["datetime"] = pd.NaT
    
    mdf_full = mdf_full.sort_values("datetime").reset_index(drop=True)
    
    # Convert price columns to numeric BEFORE ATR calculation
    for col in ['high', 'low', 'close', 'open']:
        if col in mdf_full.columns:
            mdf_full[col] = pd.to_numeric(mdf_full[col], errors='coerce')
    
    # Calculate rolling ATR on FULL history (no look-ahead bias)
    mdf_full = add_rolling_atr(mdf_full)
    
    # ===============================
    # Step B: Get ATR at Entry (from candle preceding buy_time)
    # ===============================
    # Build the entry datetime for the trade
    if date_str and buy_time:
        try:
            entry_datetime = pd.to_datetime(f"{date_str} {buy_time.strftime('%H:%M:%S')}")
            # Match timezone if needed (data may be tz-aware)
            if mdf_full['datetime'].dt.tz is not None:
                entry_datetime = entry_datetime.tz_localize(mdf_full['datetime'].dt.tz)
        except:
            entry_datetime = None
    else:
        entry_datetime = None
    
    # Get ATR from the candle immediately preceding entry time
    atr_value = None
    if entry_datetime is not None:
        pre_entry_candles = mdf_full[mdf_full['datetime'] < entry_datetime]
        if len(pre_entry_candles) > 0:
            atr_value = pre_entry_candles.iloc[-1]['ATR']
            if pd.isna(atr_value):
                atr_value = None
    
    # ===============================
    # v5.0 CONSISTENT GEOMETRY: Use Phase-3 Levels Exactly
    # ===============================
    # Pillar 1: Phase 4 is a "Slave" to Phase 3
    # We use Entry, Stop, Target from Phase 3 without modification
    
    # Risk_R is pre-calculated in the dataframe from Phase 3 geometry
    risk_r = row.get("risk_r", buy_price - stop)
    
    # Validate geometry (should already be validated, but double-check)
    if risk_r <= 0:
        print(f"⚠️ {symbol}: Invalid geometry (R={risk_r:.4f})")
        risk_r = buy_price * 0.015  # Fallback to 1.5% (should never happen)
    
    # Use Phase-3 target exactly (already bound above from row)
    # target is already set from row["target_price"]
    
    # Initialize trailing stop at initial Phase-3 stop
    current_stop = stop  # Start at the Phase-3 stop
    initial_stop = stop  # Store for reference
    highest_price = buy_price  # Track MFE

    
    # ===============================
    # Filter for trade date (for exit simulation)
    # ===============================
    if date_str:
        mdf = mdf_full[mdf_full["datetime"].dt.strftime("%Y-%m-%d") == date_str].copy()
    else:
        mdf = mdf_full.copy()
    
    # ===============================
    # Calculate when High and Low occurred (for analysis)
    # ===============================
    high_time = ""
    low_time = ""
    if len(mdf) > 0 and 'high' in mdf.columns and 'low' in mdf.columns:
        try:
            # Find the time when the day high occurred
            high_idx = mdf['high'].idxmax()
            if not pd.isna(high_idx):
                high_time = mdf.loc[high_idx, 'datetime'].strftime('%H:%M') if pd.notna(mdf.loc[high_idx, 'datetime']) else ""
            
            # Find the time when the day low occurred
            low_idx = mdf['low'].idxmin()
            if not pd.isna(low_idx):
                low_time = mdf.loc[low_idx, 'datetime'].strftime('%H:%M') if pd.notna(mdf.loc[low_idx, 'datetime']) else ""
        except:
            pass

    if True:  # Always true now since we checked file exists above
        # mdf is now filtered for trade date

        # Find the actual 1-minute entry time
        # Phase-3 entry time is when the 5-min candle closed (e.g., 09:35:00)
        # We need to find which 1-min candle within that 5-min period hit the entry price
        # Look backwards from buy_time to (buy_time - 5 minutes)
        actual_entry_time = None
        
        # Calculate the start of the 5-minute window
        buy_hour = buy_time.hour
        buy_minute = buy_time.minute
        
        # Round down to nearest 5-minute interval
        window_start_minute = (buy_minute // 5) * 5
        window_start = time(buy_hour, window_start_minute)
        
        for _, m in mdf.iterrows():
            if pd.isna(m.get("datetime")):
                continue
            t = m["datetime"].time()
            
            # Look within the 5-minute window ending at buy_time
            if t < window_start or t > buy_time:
                continue
                
            high = float(m.get("high", np.nan))
            # Entry happens when High >= Entry Price (trigger hit)
            if not np.isnan(high) and high >= buy_price:
                actual_entry_time = t
                break
        
        # If we found a precise entry time, use it; otherwise use Phase-3 time
        if actual_entry_time:
            buy_time = actual_entry_time

        for candle_idx, (_, m) in enumerate(mdf.iterrows()):
            # skip rows without valid datetime
            if pd.isna(m.get("datetime")):
                continue
            t = m["datetime"].time()

            # start checking only from the candle strictly after the buy_time
            if t <= buy_time:
                continue

            had_post_buy_candle = True
            last_candle_close = float(m.get("close", np.nan)) if not pd.isna(m.get("close", np.nan)) else None
            last_candle_time = t

            high = float(m.get("high", np.nan))
            low = float(m.get("low", np.nan))
            op = float(m.get("open", np.nan))
            close = float(m.get("close", np.nan))
            
            # v5.0: Track highest price for MFE calculation
            if not np.isnan(high) and high > highest_price:
                highest_price = high
            
            # Calculate current MFE in R units
            mfe = highest_price - buy_price
            mfe_r = mfe / risk_r if risk_r > 0 else 0
            
            # v5.0 MODULE 4: Update trailing stop via Half-Risk Ladder
            current_stop = calculate_half_risk_ladder(buy_price, risk_r, mfe_r, current_stop)
            
            # v5.0 MODULE 3: Calculate minutes in trade for No-Progress check
            # Convert times to minutes since midnight for calculation
            entry_minutes = buy_time.hour * 60 + buy_time.minute
            current_minutes = t.hour * 60 + t.minute
            minutes_in_trade = current_minutes - entry_minutes
            
            # Check for No-Progress (stagnation)
            should_exit_no_progress, no_progress_reason = check_no_progress(minutes_in_trade, mfe_r, mode_letter)
            if should_exit_no_progress:
                sell_price = close if not np.isnan(close) else buy_price
                sell_time = t
                exit_reason = no_progress_reason
                break
            
            # v5.0: Check current (possibly trailed) stop
            hit_stop = (not np.isnan(low)) and (low <= current_stop)
            hit_target = (not np.isnan(high)) and (high >= target)

            # Both hit in same candle → approximate intrabar order using distance from open
            if hit_stop and hit_target:
                dist_stop = abs(op - current_stop) if not np.isnan(op) else float('inf')
                dist_target = abs(op - target) if not np.isnan(op) else float('inf')
                if dist_stop <= dist_target:
                    sell_price = current_stop
                    exit_reason = "STOP_LOSS"
                else:
                    sell_price = target
                    exit_reason = "TARGET_HIT"
                sell_time = t
                break

            if hit_stop:
                sell_price = current_stop
                sell_time = t
                exit_reason = "STOP_LOSS"
                break

            if hit_target:
                sell_price = target
                sell_time = t
                exit_reason = "TARGET_HIT"
                break

            # Force-exit at or after FORCE_EXIT_TIME (TIME_EOD)
            if t >= FORCE_EXIT_TIME:
                sell_price = close if not np.isnan(close) else m.get("close", buy_price)
                sell_time = t
                exit_reason = "TIME_EOD"
                break


        # If no explicit exit was found but there were candles after the buy_time,
        # close at the last available candle close to approximate end-of-data exit.
        if exit_reason == "NO_EXIT" and had_post_buy_candle and last_candle_close is not None:
            sell_price = last_candle_close
            sell_time = last_candle_time
            exit_reason = "NO_EXIT_LASTCANDLE"

    # ===============================
    # v5.0 MODULE 6: REALISTIC FRICTION MODELING
    # ===============================
    # Slippage: Buy at Price × 1.001, Sell at Price × 0.999
    adjusted_entry = buy_price * (1 + SLIPPAGE_PCT)
    adjusted_exit = sell_price * (1 - SLIPPAGE_PCT)
    
    # Transaction costs on turnover
    turnover = (adjusted_entry + adjusted_exit) * qty
    txn_cost = turnover * TRANSACTION_COST_PCT
    
    # Net P&L with friction
    pnl = ((adjusted_exit - adjusted_entry) * qty) - txn_cost
    
    # Update daily P&L tracker for Kill Switch
    if date_str:
        daily_pnl_tracker[date_str] = daily_pnl_tracker.get(date_str, 0) + pnl

    # Normalize sell_time to string for consistent Excel output
    st = ""
    try:
        if isinstance(sell_time, time):
            st = sell_time.strftime("%H:%M:%S")
        elif isinstance(sell_time, str):
            st = sell_time
        elif pd.isna(sell_time) or sell_time is None:
            st = ""
        else:
            st = str(sell_time)
    except Exception:
        st = ""

    sell_prices.append(sell_price)
    sell_times.append(st)
    exit_reasons.append(exit_reason)
    pnls.append(pnl)
    txn_cost_list.append(txn_cost)  # v3.5: Track actual transaction cost
    risk_r_list.append(risk_r)
    stop_dynamic_list.append(current_stop)  # Track final stop (may be trailed)
    target_dynamic_list.append(target)
    high_time_list.append(high_time)
    low_time_list.append(low_time)
    qty_executed_list.append(qty)  # v3.5: Track executed qty


df["buy_price"] = df["entry_price"]
df["sell_price"] = sell_prices
df["sell_time"] = sell_times
df["exit_reason"] = exit_reasons
df["pnl"] = pnls
df["txn_cost"] = txn_cost_list  # v3.5: Actual transaction cost from engine
df["risk_r"] = risk_r_list
df["stop_dynamic"] = stop_dynamic_list
df["target_dynamic"] = target_dynamic_list
df["high_time"] = high_time_list
df["low_time"] = low_time_list
df["qty_executed"] = qty_executed_list  # v3.5: Actual executed qty


# Ensure integer quantities
df["quantity"] = df["quantity"].astype(int)

# ===============================
# FINAL OUTPUT: Build 4 sheets
# Sheets: Trade Log, Daily Summary, Performance, Algorithm Config
# ===============================

# Prepare Trade Log with requested columns (use existing df columns where available)
trade_log_cols = [
    'Date', 'Stock', 'Mode', 'Weight', 
    'Open', 'High', 'HighTime', 'Low', 'LowTime', 'Close', 'Volume', 'DataSource', 'CandleTimeframe',
    'EntryTime', 'EntryPrice', 'StopLoss', 'Target', 'ExitTime',
    'ExitReason', 'ExitPrice', 'Quantity', 'InvestedAmount', 'ProfitBeforeCosts',
    'TransactionCost', 'FinalProfit', 'P&L%'
]

# Build trade log rows
trade_rows = []
for _, r in df.iterrows():
    date_val = r.get('date')
    try:
        date_str = pd.to_datetime(date_val).strftime('%Y-%m-%d')
    except Exception:
        date_str = str(date_val)

    invested = float(r.get('trade_value', 0.0))
    exit_reason = r.get('exit_reason', '')
    qty_exec = r.get('qty_executed', 0)
    
    # v3.5 FIX A: P&L from engine is ALREADY NET of friction
    # Do NOT subtract costs again - that was the double-charging bug
    if qty_exec == 0:  # TRUE SKIP: no execution
        tx_cost = 0
        final_profit = 0
        invested = 0
        pl_pct = 0
    else:
        tx_cost = float(r.get('txn_cost', 0.0))  # Use actual cost from engine
        final_profit = float(r.get('pnl', 0.0))  # This is ALREADY net
        pl_pct = (final_profit / invested * 100.0) if invested > 0 else 0.0


    row = {
        'Date': date_str,
        'Stock': r.get('symbol'),
        'Mode': r.get('mode', ''),
        'Weight': float(r.get('weight', 0.0)),
        'Open': np.nan,
        'High': np.nan,
        'HighTime': r.get('high_time', ''),  # Time when day high occurred
        'Low': np.nan,
        'LowTime': r.get('low_time', ''),    # Time when day low occurred
        'Close': np.nan,
        'Volume': np.nan,
        'DataSource': r.get('DataSource') or ONE_MIN_DATA_DIR,
        'CandleTimeframe': r.get('CandleTimeframe') or '1m',
        'EntryTime': r.get('buy_time'),
        'EntryPrice': float(r.get('buy_price', 0.0)),
        'StopLoss': float(r.get('stop_price', np.nan)),
        'Target': float(r.get('target_price', np.nan)),
        'ExitTime': r.get('sell_time'),
        'ExitReason': r.get('exit_reason'),
        'ExitPrice': float(r.get('sell_price', 0.0)),
        'Quantity': int(r.get('quantity', 0)),
        'InvestedAmount': invested,
        'ProfitBeforeCosts': final_profit + tx_cost,  # Gross = Net + Costs
        'TransactionCost': tx_cost,
        'FinalProfit': final_profit,
        'P&L%': pl_pct
    }
    trade_rows.append(row)

trade_log_df = pd.DataFrame(trade_rows, columns=trade_log_cols)

ohlcv_path = os.path.join('downloaded_data', 'daily_candles_nifty500.xlsx')
if os.path.exists(ohlcv_path):
    try:
        ohl = pd.read_excel(ohlcv_path)
        ohl.columns = [str(c).strip() for c in ohl.columns]
        
        # Normalize symbol column
        if 'Symbol' in ohl.columns:
            ohl['Symbol'] = ohl['Symbol'].astype(str).str.upper().str.strip()
        elif 'symbol' in ohl.columns:
            ohl['Symbol'] = ohl['symbol'].astype(str).str.upper().str.strip()
        
        # Handle Datetime column (not Date)
        if 'Datetime' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['Datetime']).dt.strftime('%Y-%m-%d')
        elif 'Date' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['Date']).dt.strftime('%Y-%m-%d')
        elif 'date' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['date']).dt.strftime('%Y-%m-%d')

        # keep only relevant OHLCV cols that exist
        ohlv_cols = [c for c in ['Open','High','Low','Close','Volume'] if c in ohl.columns]

        # create lookup dict keyed by (symbol,date)
        ohl['__key'] = ohl['Symbol'].astype(str) + '||' + ohl['Date'].astype(str)
        ohl_lookup = ohl.set_index('__key')[ohlv_cols].to_dict(orient='index')

        # normalize trade keys and attempt fast lookup
        def try_get_ohlv(stock, date):
            if pd.isna(stock) or pd.isna(date):
                return None
            k = str(stock).upper().strip() + '||' + str(date)
            if k in ohl_lookup:
                return ohl_lookup[k]
            # fallback: try exact match in symbols for that date
            for idx,row in ohl[ohl['Date']==str(date)].iterrows():
                sym = row['Symbol']
                if sym == str(stock).upper().strip():
                    return row[ohlv_cols].to_dict()
            # try partial match
            for idx,row in ohl[ohl['Date']==str(date)].iterrows():
                sym = row['Symbol']
                if str(stock).upper().strip() in sym or sym in str(stock).upper().strip():
                    return row[ohlv_cols].to_dict()
            return None

        # apply per-row fill
        for i, tr in trade_log_df.iterrows():
            stock = tr['Stock']
            datev = tr['Date']
            ovals = try_get_ohlv(stock, datev)
            if ovals:
                for c in ohlv_cols:
                    trade_log_df.at[i, c] = ovals.get(c)
        
        # Ensure OHLCV columns are numeric (not strings)
        for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
            if col in trade_log_df.columns:
                trade_log_df[col] = pd.to_numeric(trade_log_df[col], errors='coerce')
                
    except Exception as e:
        print(f"⚠️ Warning: Could not load OHLCV data: {e}")
        pass

# Ensure Bias is readable (only if column exists)
if 'Bias' in trade_log_df.columns:
    trade_log_df['Bias'] = trade_log_df['Bias'].fillna('N/A')

# Daily summary (requested columns)
daily = trade_log_df.groupby('Date').agg(
    TotalTrades=('Stock', 'count'),
    CapitalStart=('InvestedAmount', lambda x: float(C_PER_DAY)),
    CapitalInvested=('InvestedAmount', 'sum'),
    CapitalEnd=('FinalProfit', lambda x: float(C_PER_DAY) + x.sum()),
    Profit=('FinalProfit', lambda x: x[x>0].sum()),
    Loss=('FinalProfit', lambda x: abs(x[x<0].sum())),
    DailyP_and_L=('FinalProfit', 'sum')
).reset_index()

# Rename DailyP_and_L -> DailyP&L and DayStatus
daily = daily.rename(columns={'DailyP_and_L': 'DailyP&L'})
daily['DayStatus'] = daily['DailyP&L'].apply(lambda v: 'PROFIT' if v>0 else ('LOSS' if v<0 else 'BREAKEVEN'))

# Reorder daily columns as requested
daily_summary_cols = ['Date','TotalTrades','CapitalStart','CapitalInvested','CapitalEnd','Profit','Loss','DailyP&L','DayStatus']
daily_summary = daily[daily_summary_cols]

# ===============================
# PERFORMANCE SHEET - CRYSTAL CLEAR FORMAT
# ===============================

# Calculate overall metrics first
total_trades = len(trade_log_df)

# Filter out skipped trades (those with zero P&L are skipped, not breakeven)
executed_trades_df = trade_log_df[~trade_log_df['ExitReason'].str.contains('SKIPPED', na=False)]
skipped_trades_df = trade_log_df[trade_log_df['ExitReason'].str.contains('SKIPPED', na=False)]
executed_count = len(executed_trades_df)
skipped_count = len(skipped_trades_df)

wins_df = executed_trades_df[executed_trades_df['FinalProfit']>0]
losses_df = executed_trades_df[executed_trades_df['FinalProfit']<0]
win_count = len(wins_df)
loss_count = len(losses_df)

# Win rate: calculated only from executed trades (excludes skipped)
win_rate = (win_count / executed_count * 100.0) if executed_count > 0 else 0.0
loss_rate = (loss_count / executed_count * 100.0) if executed_count > 0 else 0.0

total_invested = executed_trades_df['InvestedAmount'].sum()
total_profit_amount = wins_df['FinalProfit'].sum() if win_count > 0 else 0.0
total_loss_amount = abs(losses_df['FinalProfit'].sum()) if loss_count > 0 else 0.0
net_pnl = executed_trades_df['FinalProfit'].sum()
roi_percent = (net_pnl / total_invested * 100.0) if total_invested > 0 else 0.0

# Calculate daily P&L for average per day metrics (only executed trades)
daily_pnl = executed_trades_df.groupby('Date')['FinalProfit'].sum()
daily_profit = daily_pnl[daily_pnl > 0]
daily_loss = daily_pnl[daily_pnl < 0]
avg_net_profit_per_day = daily_profit.mean() if len(daily_profit) > 0 else 0.0
avg_net_loss_per_day = abs(daily_loss.mean()) if len(daily_loss) > 0 else 0.0

avg_win = wins_df['FinalProfit'].mean() if win_count > 0 else 0.0
avg_loss = abs(losses_df['FinalProfit'].mean()) if loss_count > 0 else 0.0

# Daily trading statistics (only executed trades)
trades_per_day = executed_trades_df.groupby('Date').size()
unique_dates = executed_trades_df['Date'].unique()
total_trading_days = len(unique_dates)
avg_scripts_per_day = trades_per_day.mean() if len(trades_per_day) > 0 else 0.0

# Min/Max scripts traded on a given day (no dates since multiple days could have same count)
if len(trades_per_day) > 0:
    min_trades_count = trades_per_day.min()
    max_trades_count = trades_per_day.max()
else:
    min_trades_count = 0
    max_trades_count = 0

# Capital invested per day statistics (only executed trades)
capital_per_day = executed_trades_df.groupby('Date')['InvestedAmount'].sum()
avg_capital_per_day = capital_per_day.mean() if len(capital_per_day) > 0 else 0.0
min_capital_per_day = capital_per_day.min() if len(capital_per_day) > 0 else 0.0
max_capital_per_day = capital_per_day.max() if len(capital_per_day) > 0 else 0.0

# Create SUMMARY section (top of sheet)
summary_data = [
    ['═══════════════════════════════════════════════════════════════', ''],
    ['                    📊 BACKTEST PERFORMANCE SUMMARY', ''],
    ['═══════════════════════════════════════════════════════════════', ''],
    ['', ''],
    ['💰 FINAL RESULT', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Capital Deployed', f'₹{total_invested:,.2f}'],
    ['Avg Net Profit/Day', f'₹{avg_net_profit_per_day:,.2f}'],
    ['Avg Net Loss/Day', f'₹{avg_net_loss_per_day:,.2f}'],
    ['Net Profit/Loss', f'₹{net_pnl:,.2f}'],
    ['Return on Investment (ROI)', f'{roi_percent:.2f}%'],
    ['Final Status', 'PROFIT ✅' if net_pnl > 0 else ('LOSS ❌' if net_pnl < 0 else 'BREAKEVEN')],
    ['', ''],
    ['📊 TRADE STATISTICS', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Trades', f'{total_trades}'],
    ['Executed Trades', f'{executed_count}'],
    ['Skipped Trades', f'{skipped_count}'],
    ['Winning Trades', f'{win_count} ({win_rate:.1f}%)'],
    ['Losing Trades', f'{loss_count} ({loss_rate:.1f}%)'],
    ['', ''],
    ['📅 DAILY TRADING STATISTICS', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Trading Days', f'{total_trading_days}'],
    ['Avg # Scripts Traded/Day', f'{avg_scripts_per_day:.2f}'],
    ['Min # Scripts Traded', f'{min_trades_count}'],
    ['Max # Scripts Traded', f'{max_trades_count}'],
    ['', ''],
    ['💵 CAPITAL STATISTICS', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Avg Capital Invested/Day', f'₹{avg_capital_per_day:,.2f}'],
    ['Min Capital Invested/Day', f'₹{min_capital_per_day:,.2f}'],
    ['Max Capital Invested/Day', f'₹{max_capital_per_day:,.2f}'],
    ['', ''],
    ['💵 AVERAGE PER TRADE', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Average Win', f'₹{avg_win:,.2f}'],
    ['Average Loss', f'₹{avg_loss:,.2f}'],
    ['Average P&L per Trade', f'₹{net_pnl/executed_count:,.2f}' if executed_count > 0 else '₹0.00'],
    ['', ''],
    ['═══════════════════════════════════════════════════════════════', ''],
]

summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])

# Use summary as the performance sheet (no PromptType breakdown needed)
performance_df = summary_df

# Algorithm Config sheet - produce key/value table from v3.5 variables
alg_config = {
    'Version': '3.5',
    'Architecture': 'Regime-Gated Execution with Consistent Geometry & Progress Logic',
    'CapitalPerDay': f"₹{C_PER_DAY:,}",
    'RiskPerTrade': f"₹{RISK_PER_TRADE:,}",
    'MaxNotional': f"₹{MAX_NOTIONAL:,}",
    'DailyLossLimit': f"₹{abs(DAILY_LOSS_LIMIT):,}",
    'CandleTimeframe': '1m',
    'DataSource': ONE_MIN_DATA_DIR,
    'RegimeModeA': '09:15-10:15',
    'RegimeModeB': '09:15-11:00, 13:30-15:00',
    'RegimeModeC': '10:00-14:00',
    'NoProgress_A': f"{NO_PROGRESS_CONFIG['A']['time_limit']}min @ {NO_PROGRESS_CONFIG['A']['threshold']}R",
    'NoProgress_B': f"{NO_PROGRESS_CONFIG['B']['time_limit']}min @ {NO_PROGRESS_CONFIG['B']['threshold']}R",
    'NoProgress_C': f"{NO_PROGRESS_CONFIG['C']['time_limit']}min @ {NO_PROGRESS_CONFIG['C']['threshold']}R",
    'ForceExitTime': FORCE_EXIT_TIME.strftime('%H:%M'),
    'SlippagePercent': f"{SLIPPAGE_PCT*100:.2f}%",
    'TransactionCostPercent': f"{TRANSACTION_COST_PCT*100:.2f}%"
}

config_df = pd.DataFrame(list(alg_config.items()), columns=['Parameter','Value'])

# Write sheets
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    trade_log_df.to_excel(writer, sheet_name='Trade Log', index=False)
    daily_summary.to_excel(writer, sheet_name='Daily Summary', index=False)
    performance_df.to_excel(writer, sheet_name='Performance', index=False)
    config_df.to_excel(writer, sheet_name='Algorithm Config', index=False)

# Load workbook and apply formatting
wb = openpyxl.load_workbook(OUTPUT_FILE)

def format_sheet(ws, header_fill=None, pnl_col_name=None):
    hdr = Font(bold=True)
    for cell in ws[1]:
        cell.font = hdr
        cell.alignment = Alignment(horizontal='center')
        if header_fill:
            cell.fill = header_fill
    if pnl_col_name:
        pnl_col = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).lower() == pnl_col_name.lower():
                pnl_col = idx
                break
        if pnl_col and ws.max_row > 1:
            col_letter = openpyxl.utils.get_column_letter(pnl_col)
            green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            range_str = f'{col_letter}2:{col_letter}{ws.max_row}'
            ws.conditional_formatting.add(range_str,
                                          CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green))
            ws.conditional_formatting.add(range_str,
                                          CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red))

# Format Trade Log with blue header
if 'Trade Log' in wb.sheetnames:
    ws = wb['Trade Log']
    blue = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    format_sheet(ws, header_fill=blue, pnl_col_name='FinalProfit')

# Format Daily Summary with bold headers and conditional coloring on DailyP&L
if 'Daily Summary' in wb.sheetnames:
    ws2 = wb['Daily Summary']
    yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    format_sheet(ws2, header_fill=yellow, pnl_col_name='DailyP&L')

# Format Performance (special handling for summary format)
if 'Performance' in wb.sheetnames:
    ws3 = wb['Performance']
    
    # Style for section headers (lines with emojis and titles)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=14, color='FFFFFF')
    
    # Style for separator lines
    separator_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Style for metric names
    metric_font = Font(bold=True, size=11)
    
    # Style for values
    value_font = Font(size=11)
    
    # Apply formatting row by row
    for row_idx in range(1, ws3.max_row + 1):
        metric_cell = ws3.cell(row=row_idx, column=1)
        value_cell = ws3.cell(row=row_idx, column=2)
        
        metric_text = str(metric_cell.value) if metric_cell.value else ''
        value_text = str(value_cell.value) if value_cell.value else ''
        
        # Header rows (with emojis or equals signs)
        if '═' in metric_text or '📊' in metric_text:
            metric_cell.font = header_font
            metric_cell.fill = header_fill
            value_cell.fill = header_fill
            # Merge cells for header rows
            if '📊' in metric_text or '═' in metric_text:
                ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                metric_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Section titles (with emojis)
        elif any(emoji in metric_text for emoji in ['💰', '📈', '📊', '💵']):
            metric_cell.font = Font(bold=True, size=12, color='1F4E78')
            metric_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            value_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        
        # Separator lines
        elif '─' in metric_text:
            metric_cell.fill = separator_fill
            value_cell.fill = separator_fill
            ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        
        # Data rows
        elif metric_text and metric_text.strip():
            metric_cell.font = metric_font
            value_cell.font = value_font
            
            # Color code profit/loss values
            if 'Net Profit/Loss' in metric_text or 'Final Status' in metric_text or 'Net Amount' in metric_text:
                if '✅' in value_text or (value_text.startswith('₹') and float(value_text.replace('₹','').replace(',','')) > 0):
                    value_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    value_cell.font = Font(bold=True, size=12, color='006100')
                elif '❌' in value_text or (value_text.startswith('₹') and float(value_text.replace('₹','').replace(',','')) < 0):
                    value_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    value_cell.font = Font(bold=True, size=12, color='9C0006')
            
            # Highlight ROI
            if 'ROI' in metric_text:
                value_cell.font = Font(bold=True, size=12)
    
    # Adjust column widths
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 25

# Format Algorithm Config (dark header, clear key/value)
if 'Algorithm Config' in wb.sheetnames:
    ws4 = wb['Algorithm Config']
    dark = PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid')
    for cell in ws4[1]:
        cell.fill = dark
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    # alternate shading for readability
    for r in range(2, ws4.max_row+1):
        fill = PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid') if r%2==0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        for c in range(1, ws4.max_column+1):
            ws4.cell(row=r, column=c).fill = fill

wb.save(OUTPUT_FILE)

print("✅ Phase-4 BACKTEST (1-minute) completed successfully")
print(f"📄 Output saved to: {OUTPUT_FILE}")
# Show net P&L (after transaction costs), excluding skipped trades
active_trades = trade_log_df[~trade_log_df['ExitReason'].str.contains('SKIPPED', na=False)]
print(f"Net P&L (after costs): ₹{active_trades['FinalProfit'].sum():,.2f}")
